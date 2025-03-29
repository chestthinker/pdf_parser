import openpyxl

def write_to_excel(sections, file_path, row_height, col_width, title):
    """Write data to Excel file with formatted styles and structure"""
    
    try:
        # Initialize workbook and worksheet
        workbook = openpyxl.Workbook()        # Create a new workbook
        worksheet = workbook.active           # Get the active worksheet

        # Add header row with titles
        sections.insert(0, title)             # Insert title row at the beginning
        num_rows = len(sections)

        # Write all data to worksheet cells
        for row_index, row_content in enumerate(sections):
            for col_index, elem in enumerate(row_content):
                worksheet.cell(row=row_index+1, column=col_index+1, value=elem)

        # Set body content styling
        font_style = openpyxl.styles.Font(name="微软雅黑", size=8, bold=False)
        alignment_style = openpyxl.styles.Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # Apply styling to body rows (excluding header)
        for row in worksheet.iter_rows(min_row=2, max_row=num_rows):
            for cell in row:
                cell.font = font_style
                cell.alignment = alignment_style

        # Special center alignment for first two columns
        alignment_style = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in worksheet.iter_rows(min_row=2, max_row=num_rows, min_col=1, max_col=2):
            row[0].alignment = alignment_style  # Column 1
            row[1].alignment = alignment_style  # Column 2

        # Set column widths from predefined sizes
        for col_index in range(len(col_width)):
            col_letter = openpyxl.utils.get_column_letter(col_index+1)
            worksheet.column_dimensions[col_letter].width = col_width[col_index]

        # Set row heights for data rows
        if row_height == 'auto':
            auto_adjust_row_height(worksheet)
        else:
            for row_index in range(1, num_rows+1):
                worksheet.row_dimensions[row_index+1].height = row_height

    
        # Configure header row
        worksheet.freeze_panes = "A2"           # Freeze header row
        worksheet.row_dimensions[1].height = 30 # Set header row height
        
        # Header styling (bold font + center alignment + light blue background)
        font_style = openpyxl.styles.Font(name="微软雅黑", size=10, bold=True)
        alignment_style = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True)
        fill_style = openpyxl.styles.PatternFill(start_color="ADD8FF", fill_type="solid")  # Light blue
        
        # Apply header styling to all cells in first row
        for cell in worksheet[1]:
            cell.font = font_style
            cell.alignment = alignment_style
            cell.fill = fill_style

        # Save and finalize
        workbook.save(file_path)
        print(f" Complete: -> {file_path} ")
        
    except Exception as e:
        print(f" Write to file <{file_path}> Error \n {e} \n")







def auto_adjust_row_height(ws):
    """Automatically adjust row heights based on wrapped text content"""
    for row in ws.iter_rows():
        max_lines = 1  # Default to at least 1 line
        for cell in row:
            if cell.alignment and cell.alignment.wrap_text and cell.value:
                # Get column width (character count estimate)
                col_letter = openpyxl.utils.get_column_letter(cell.column)
                col_width = ws.column_dimensions[col_letter].width
                
                # Get font size (default to 11 if not set)
                font_size = cell.font.size if cell.font else 11
                
                # Calculate required lines (assuming 1 width unit per Chinese character)
                text = str(cell.value)
                lines = text.split('\n')  # Handle manual line breaks
                total_lines = 0
                
                for line in lines:
                    line_length = len(line)
                    # Calculate lines needed for automatic wrapping
                    lines_needed = (line_length // col_width) + 1 if line_length % col_width else line_length // col_width
                    total_lines += lines_needed
                
                max_lines = max(max_lines, total_lines)
        
        # Set row height (1.1x font size per line, clamped between 15-60)
        if max_lines > 1:
            line_height = font_size * 1.1 * max_lines
            line_height = min(80, max(15, line_height))  # Constrain height range
            ws.row_dimensions[row[0].row].height = line_height

